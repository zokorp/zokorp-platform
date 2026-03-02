#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

URL_RE = re.compile(r"https?://[^\s\]\)>\"']+")


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = value.replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value or "item"


def normalize_header(value: str) -> str:
    value = value.replace("\xa0", " ")
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value)


def detect_header_row(ws) -> int:
    best_row = 1
    best_count = -1
    for row in range(1, min(ws.max_row, 8) + 1):
        count = sum(
            1
            for col in range(1, ws.max_column + 1)
            if ws.cell(row, col).value not in (None, "")
        )
        if count > best_count:
            best_count = count
            best_row = row
    return best_row


def extract_urls(text: str) -> list[str]:
    if not text:
        return []
    return [match.group(0).rstrip(".,;") for match in URL_RE.finditer(text)]


def cell_text_and_link(cell) -> tuple[str, str | None]:
    if cell.value is None:
        return "", None

    text = str(cell.value).replace("\xa0", " ").strip()
    link = None
    if cell.hyperlink and cell.hyperlink.target:
        link = str(cell.hyperlink.target).strip()

    if not link:
        urls = extract_urls(text)
        if urls:
            link = urls[0]

    return text, link


def find_column(header_map: dict[str, int], *patterns: str) -> int | None:
    for pattern in patterns:
        pattern = normalize_header(pattern)
        for header, col in header_map.items():
            if pattern in header:
                return col
    return None


def get_value(row_data: dict[str, Any], key: str) -> str:
    value = row_data.get(key)
    if value is None:
        return ""
    return str(value).strip()


def build_rows(ws):
    header_row = detect_header_row(ws)
    headers: dict[int, str] = {}
    header_map: dict[str, int] = {}

    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if value is None:
            continue
        text = str(value).replace("\xa0", " ").strip()
        if not text:
            continue
        headers[col] = text
        header_map[normalize_header(text)] = col

    rows: list[dict[str, Any]] = []
    for row in range(header_row + 1, ws.max_row + 1):
        row_texts: dict[str, str] = {}
        row_links: dict[str, str] = {}

        for col, header in headers.items():
            text, link = cell_text_and_link(ws.cell(row, col))
            if text:
                row_texts[header] = text
            if link:
                row_links[header] = link

        if not row_texts and not row_links:
            continue

        rows.append(
            {
                "row_number": row,
                "values": row_texts,
                "links": row_links,
            }
        )

    return header_row, header_map, rows


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def write_json(path: Path, payload: Any) -> None:
    ensure_dir(path.parent)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")


def write_readme(path: Path, title: str, sections: dict[str, Any]) -> None:
    lines = [f"# {title}", ""]
    for key, value in sections.items():
        lines.append(f"## {key}")
        if isinstance(value, list):
            for item in value:
                lines.append(f"- {item}")
        else:
            lines.append(str(value))
        lines.append("")
    path.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")


def write_link_file(path: Path, url: str | None) -> None:
    ensure_dir(path.parent)
    value = (url or "").strip() or "NOT_AVAILABLE"
    path.write_text(value + "\n", encoding="utf-8")


def parse_sv_designations(rows: list[dict[str, Any]]):
    grouped: dict[str, dict[str, Any]] = {}

    for row in rows:
        values = row["values"]
        links = row["links"]

        service = values.get("Service", "").strip()
        if not service:
            continue

        checklist_url = links.get("Archive_ Validation Checklist") or ""
        if not checklist_url:
            checklist_text = values.get("Archive_ Validation Checklist", "")
            urls = extract_urls(checklist_text)
            checklist_url = urls[0] if urls else ""

        if not checklist_url:
            continue

        key = service
        if key not in grouped:
            grouped[key] = {
                "service": service,
                "service_category": values.get(
                    'Service Category  (updated on April 18, 2019 based on "Product" tab of AWS homepage)',
                    values.get('Service Category\xa0 (updated on April 18, 2019 based on "Product" tab of AWS homepage)', ""),
                ).strip(),
                "team_alias": values.get("Archive_Service Team Alias", "").strip(),
                "designation_webpage": links.get("Designation Webpage")
                or (extract_urls(values.get("Designation Webpage", "")) or [""])[0],
                "versions": [],
            }

        grouped[key]["versions"].append(
            {
                "source_row": row["row_number"],
                "status": values.get("Status", "").strip(),
                "delivery_count": values.get("Delivery Count", "").strip(),
                "ready_count": values.get("Ready Count", "").strip(),
                "checklist_url": checklist_url,
                "notes": values.get("Archive_PSA who supported VCL build v1", "").strip(),
            }
        )

    items = []
    for _, item in sorted(grouped.items(), key=lambda kv: kv[0].lower()):
        seen = set()
        deduped = []
        for version in item["versions"]:
            url = version["checklist_url"]
            if url in seen:
                continue
            seen.add(url)
            deduped.append(version)
        item["versions"] = deduped
        items.append(item)

    return items


def parse_service_specializations(rows: list[dict[str, Any]]):
    records = []
    for row in rows:
        values = row["values"]
        links = row["links"]

        designation = values.get("AWS Service Designation", "").strip()
        if not designation:
            continue

        checklist_url = links.get("Validation Checklist (S3 link)") or ""
        if not checklist_url:
            checklist_url = (extract_urls(values.get("Validation Checklist (S3 link)", "")) or [""])[0]

        guide_url = links.get("VCL Calibration  Guide") or ""
        if not guide_url:
            guide_url = (extract_urls(values.get("VCL Calibration  Guide", "")) or [""])[0]

        partner_type = values.get("Partner Type/Path", "").strip()
        partner_type_lower = partner_type.lower()

        tracks = []
        if "service" in partner_type_lower:
            tracks.append("sdp")
        if "software" in partner_type_lower or "isv" in partner_type_lower:
            tracks.append("srp")
        if not tracks:
            tracks.append("unclassified")

        records.append(
            {
                "source_row": row["row_number"],
                "designation_id": values.get("Designation ID (SFDC)", "").strip(),
                "designation": designation,
                "domain": values.get("Domain", "").strip(),
                "business_domain_leader": values.get("Business Domain Leader", "").strip(),
                "partner_type_path": partner_type,
                "accepting_applications": values.get("Accepting Applications?", "").strip(),
                "business_owner": values.get("Designation (Business) Owner", "").strip(),
                "business_owner_email": values.get("Designation (Business) Owner email", "").strip(),
                "technical_owner": values.get("Technical Owner", "").strip(),
                "technical_owner_email": values.get("Technical Owner Email", "").strip(),
                "technical_validation_administrator": values.get("Technical Validation Adminstrator", "").strip(),
                "program_pm": values.get("Program Onboarding PM", "").strip(),
                "program_pm_email": values.get("Program Onboarding PM Email", "").strip(),
                "marketing_contact": values.get("Marketing Contact", "").strip(),
                "sfdc_designation": links.get("SFDC Designation")
                or (extract_urls(values.get("SFDC Designation", "")) or [""])[0],
                "checklist_url": checklist_url,
                "calibration_guide_url": guide_url,
                "tracks": tracks,
            }
        )

    return records


def parse_competencies(rows: list[dict[str, Any]]):
    records = []
    for row in rows:
        values = row["values"]
        links = row["links"]

        designation = values.get("Competency Designation", "").strip()
        if not designation:
            continue

        checklist_url = links.get("Validation Checklist") or ""
        if not checklist_url:
            checklist_url = (extract_urls(values.get("Validation Checklist", "")) or [""])[0]

        guide_url = links.get("VCL Calibration Guide") or ""
        if not guide_url:
            guide_url = (extract_urls(values.get("VCL Calibration Guide", "")) or [""])[0]

        partner_type = values.get("Partner Type/Path", "").strip()

        records.append(
            {
                "source_row": row["row_number"],
                "designation_id": values.get("Designation ID (SFDC)", "").strip(),
                "competency_designation": designation,
                "domain_title": values.get("Domain Title", "").strip(),
                "partner_type_path": partner_type,
                "accepting_applications": values.get("Accepting Applications?", "").strip(),
                "business_domain_leader": values.get("Business Domain Leader", "").strip(),
                "business_owner": values.get("Designation (Business) Owner", "").strip(),
                "business_owner_email": values.get("Designation (Business) Owner Email", "").strip(),
                "technical_owner": values.get("Technical Owner", "").strip(),
                "technical_owner_email": values.get("Technical Owner email", "").strip(),
                "technical_validation_administrator": values.get("Technical Validation Administrator", "").strip(),
                "program_pm": values.get("Program Onboarding PM", "").strip(),
                "program_pm_email": values.get("Program Onboarding PM Email", "").strip(),
                "marketing_contact": values.get("Marketing Contact", "").strip(),
                "webpage": links.get("Webpage")
                or (extract_urls(values.get("Webpage", "")) or [""])[0],
                "sfdc_designation": links.get("SFDC Designation (Use the exact Designation title in SFDC)")
                or (
                    extract_urls(
                        values.get("SFDC Designation (Use the exact Designation title in SFDC)", "")
                    )
                    or [""]
                )[0],
                "checklist_url": checklist_url,
                "calibration_guide_url": guide_url,
            }
        )

    return records


def write_service_track_folders(base_dir: Path, records: list[dict[str, Any]], track: str):
    track_dir = base_dir / track
    ensure_dir(track_dir)

    selected = [r for r in records if track in r["tracks"]]
    selected.sort(key=lambda r: (r["domain"], r["designation"]))

    for idx, record in enumerate(selected, 1):
        designation_id = record["designation_id"] or f"row-{record['source_row']}"
        folder = f"{slugify(record['designation'])}--{slugify(designation_id)}"
        folder_dir = track_dir / folder
        ensure_dir(folder_dir)

        metadata = {
            **record,
            "track": track,
            "folder": str(folder_dir.relative_to(base_dir.parent)),
        }
        write_json(folder_dir / "metadata.json", metadata)
        write_link_file(folder_dir / "resources" / "checklist-link.txt", record["checklist_url"])
        write_link_file(folder_dir / "resources" / "calibration-guide-link.txt", record["calibration_guide_url"])
        write_link_file(folder_dir / "resources" / "sfdc-designation-link.txt", record["sfdc_designation"])

        write_readme(
            folder_dir / "README.md",
            record["designation"],
            {
                "Track": track.upper(),
                "Domain": record["domain"] or "Unknown",
                "Partner Type": record["partner_type_path"] or "Unknown",
                "Checklist URL": record["checklist_url"] or "Not provided",
                "Calibration Guide URL": record["calibration_guide_url"] or "Not provided",
                "Business Owner": f"{record['business_owner']} ({record['business_owner_email']})".strip(),
                "Technical Owner": f"{record['technical_owner']} ({record['technical_owner_email']})".strip(),
                "Validation Administrator": record["technical_validation_administrator"] or "Not provided",
                "Program PM": f"{record['program_pm']} ({record['program_pm_email']})".strip(),
                "SFDC Designation": record["sfdc_designation"] or "Not provided",
            },
        )

    write_json(track_dir / "index.json", selected)


def write_competency_folders(base_dir: Path, records: list[dict[str, Any]]):
    comp_dir = base_dir / "competency"
    ensure_dir(comp_dir)

    records = sorted(records, key=lambda r: (r["domain_title"], r["competency_designation"], r["partner_type_path"]))

    for record in records:
        subtype = "services" if "service" in record["partner_type_path"].lower() else "software" if "software" in record["partner_type_path"].lower() else "other"
        comp_slug = slugify(record["competency_designation"])
        designation_id = record["designation_id"] or f"row-{record['source_row']}"

        folder_dir = comp_dir / comp_slug / subtype / slugify(designation_id)
        ensure_dir(folder_dir)

        metadata = {
            **record,
            "track": "competency",
            "subtype": subtype,
            "folder": str(folder_dir.relative_to(base_dir.parent)),
        }
        write_json(folder_dir / "metadata.json", metadata)
        write_link_file(folder_dir / "resources" / "checklist-link.txt", record["checklist_url"])
        write_link_file(folder_dir / "resources" / "calibration-guide-link.txt", record["calibration_guide_url"])
        write_link_file(folder_dir / "resources" / "webpage-link.txt", record["webpage"])
        write_link_file(folder_dir / "resources" / "sfdc-designation-link.txt", record["sfdc_designation"])

        write_readme(
            folder_dir / "README.md",
            f"{record['competency_designation']} ({subtype})",
            {
                "Track": "COMPETENCY",
                "Domain": record["domain_title"] or "Unknown",
                "Partner Type": record["partner_type_path"] or "Unknown",
                "Checklist URL": record["checklist_url"] or "Not provided",
                "Calibration Guide URL": record["calibration_guide_url"] or "Not provided",
                "Business Owner": f"{record['business_owner']} ({record['business_owner_email']})".strip(),
                "Technical Owner": f"{record['technical_owner']} ({record['technical_owner_email']})".strip(),
                "Validation Administrator": record["technical_validation_administrator"] or "Not provided",
                "Program PM": f"{record['program_pm']} ({record['program_pm_email']})".strip(),
                "SFDC Designation": record["sfdc_designation"] or "Not provided",
                "Public Webpage": record["webpage"] or "Not provided",
            },
        )

    write_json(comp_dir / "index.json", records)


def write_ftr_folders(base_dir: Path, ftr_items: list[dict[str, Any]]):
    ftr_dir = base_dir / "ftr"
    versions_dir = ftr_dir / "versions"
    ensure_dir(versions_dir)

    for item in ftr_items:
        service_slug = slugify(item["service"])
        folder_dir = versions_dir / service_slug
        ensure_dir(folder_dir)

        write_json(folder_dir / "metadata.json", item)
        write_link_file(folder_dir / "resources" / "designation-webpage-link.txt", item["designation_webpage"])

        checklist_versions_dir = folder_dir / "resources" / "checklist-versions"
        ensure_dir(checklist_versions_dir)
        for version in item["versions"]:
            row_number = version["source_row"]
            filename = f"row-{row_number}.txt"
            write_link_file(checklist_versions_dir / filename, version["checklist_url"])

        version_lines = [
            f"row {v['source_row']}: {v['checklist_url']}"
            for v in item["versions"]
        ]

        write_readme(
            folder_dir / "README.md",
            item["service"],
            {
                "Track": "FTR",
                "Service Category": item["service_category"] or "Unknown",
                "Designation Webpage": item["designation_webpage"] or "Not provided",
                "Service Team Alias": item["team_alias"] or "Not provided",
                "Checklist Versions": version_lines,
            },
        )

    write_json(ftr_dir / "index.json", ftr_items)


def export_raw_sheet_rows(ws, output_path: Path):
    header_row = detect_header_row(ws)
    headers = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        headers.append(str(value).replace("\xa0", " ").strip() if value is not None else "")

    rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        record = {}
        for col, header in enumerate(headers, 1):
            if not header:
                continue
            text, link = cell_text_and_link(ws.cell(row, col))
            if text:
                record[header] = text
            if link:
                record[f"{header}__link"] = link
        if record:
            record["__row_number"] = row
            rows.append(record)

    write_json(output_path, rows)


def build_library(source_xlsx: Path, output_root: Path, clear_output: bool):
    if clear_output and output_root.exists():
        shutil.rmtree(output_root)
    ensure_dir(output_root)

    workbook = load_workbook(source_xlsx, data_only=False)

    sv_sheet = workbook["SV Designations"]
    spec_sheet = workbook["AWS Service Specializations"]
    comp_sheet = workbook["Competencies"]

    _, _, sv_rows = build_rows(sv_sheet)
    _, _, spec_rows = build_rows(spec_sheet)
    _, _, comp_rows = build_rows(comp_sheet)

    ftr_items = parse_sv_designations(sv_rows)
    service_records = parse_service_specializations(spec_rows)
    competency_records = parse_competencies(comp_rows)

    write_ftr_folders(output_root, ftr_items)
    write_service_track_folders(output_root, service_records, "sdp")
    write_service_track_folders(output_root, service_records, "srp")
    write_competency_folders(output_root, competency_records)

    context_raw_dir = output_root / "context" / "raw_sheets"
    ensure_dir(context_raw_dir)
    for sheet_name in workbook.sheetnames:
        export_raw_sheet_rows(workbook[sheet_name], context_raw_dir / f"{slugify(sheet_name)}.json")

    catalog = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "source_workbook": str(source_xlsx),
        "counts": {
            "ftr_services": len(ftr_items),
            "sdp_designations": len([r for r in service_records if "sdp" in r["tracks"]]),
            "srp_designations": len([r for r in service_records if "srp" in r["tracks"]]),
            "competency_designations": len(competency_records),
        },
        "notes": [
            "This dataset is generated from workbook links and metadata; it does not require LLM inference.",
            "Checklist and calibration links may require AWS Partner Central authentication.",
            "Use this as deterministic source-of-truth for validator rulepack design.",
        ],
    }
    write_json(output_root / "catalog.json", catalog)

    write_readme(
        output_root / "README.md",
        "ZoKorp Validator Knowledge Library",
        {
            "Source Workbook": str(source_xlsx),
            "Generated At (UTC)": catalog["generated_at_utc"],
            "FTR Services": catalog["counts"]["ftr_services"],
            "SDP Designations": catalog["counts"]["sdp_designations"],
            "SRP Designations": catalog["counts"]["srp_designations"],
            "Competency Designations": catalog["counts"]["competency_designations"],
            "Important": [
                "Folder structure is deterministic and LLM-free.",
                "Use metadata.json in each folder for checklist and calibration links.",
                "Context/raw_sheets includes additional workbook context for future scoring logic.",
            ],
        },
    )


def main():
    parser = argparse.ArgumentParser(description="Build validator folder structure and indexes from AWS workbook.")
    parser.add_argument(
        "--source",
        default="/Users/zohaibkhawaja/Downloads/AWS Specialization Owners, Regional Leads, Resources copy.xlsx",
        help="Path to source workbook.",
    )
    parser.add_argument(
        "--output",
        default="data/validator/library",
        help="Output directory for generated folder structure.",
    )
    parser.add_argument(
        "--clear-output",
        action="store_true",
        help="Delete output directory before generation.",
    )
    args = parser.parse_args()

    source = Path(args.source).expanduser().resolve()
    output = Path(args.output).resolve()

    if not source.exists():
        raise SystemExit(f"Source workbook not found: {source}")

    build_library(source, output, args.clear_output)
    print(f"Generated validator library at: {output}")


if __name__ == "__main__":
    main()
